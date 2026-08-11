from __future__ import annotations

import io
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image,
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from platform_core.database import connect
from platform_core.projects import get_project
from project_intelligence import consolidation_service, knowledge_service, review_workflow


PAGE_WIDTH, PAGE_HEIGHT = A4


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value or "")).strip("_")
    return cleaned or "report"


def _json(value: str | None, default):
    try:
        return json.loads(value or "")
    except (TypeError, json.JSONDecodeError):
        return default


def _latest_layout_rows(project_id: int):
    """Load report data from the current layout-review schema.

    The production schema uses ``derived_json`` and ``reviewed_json``.
    Report generation deliberately uses those physical column names without
    legacy aliases, preventing SQLite from resolving an obsolete column.
    """
    with connect() as connection:
        columns = {
            row["name"]
            for row in connection.execute(
                "PRAGMA table_info(layout_review_state)"
            ).fetchall()
        }
        required = {
            "lifecycle_status",
            "confirmed_north_orientation",
            "derived_json",
            "reviewed_json",
        }
        missing = sorted(required - columns)
        if missing:
            raise RuntimeError(
                "The layout review database schema is incomplete. "
                "Missing column(s): " + ", ".join(missing)
            )

        return connection.execute(
            """
            SELECT pl.*,
                   rs.lifecycle_status,
                   rs.confirmed_north_orientation,
                   rs.derived_json,
                   rs.reviewed_json,
                   pa.analysis_version,
                   pa.vastu_score,
                   pa.overall_score,
                   pa.confidence_value,
                   pa.grade,
                   pa.strengths_json,
                   pa.cautions_json,
                   pa.findings_json,
                   pa.recommendations_json,
                   pa.result_json_path
            FROM project_layouts pl
            LEFT JOIN layout_review_state rs ON rs.layout_id=pl.id
            LEFT JOIN project_layout_analyses pa ON pa.id=(
                SELECT pa2.id
                FROM project_layout_analyses pa2
                WHERE pa2.layout_id=pl.id
                ORDER BY
                  CASE WHEN pa2.analysis_version='PROFESSIONAL-REVIEWED-1.0'
                       THEN 0 ELSE 1 END,
                  pa2.id DESC
                LIMIT 1
            )
            WHERE pl.project_id=?
            ORDER BY pl.tower,pl.floor,pl.flat_number,pl.id
            """,
            (int(project_id),),
        ).fetchall()

def _layout_payload(row) -> dict[str, Any]:
    reviewed = _json(row["reviewed_json"], {})
    derived = _json(row["derived_json"], {})
    final_directions = dict(derived)
    final_directions.update(reviewed)

    stored = {}
    result_value = str(row["result_json_path"] or "").strip()
    result_path = Path(result_value) if result_value else None
    if result_path and result_path.exists() and result_path.is_file():
        try:
            stored = json.loads(result_path.read_text(encoding="utf-8"))
        except (OSError, PermissionError, json.JSONDecodeError):
            stored = {}

    knowledge = stored.get("knowledge_assessment")
    if not knowledge:
        latest = knowledge_service.latest_assessment(int(row["id"]))
        knowledge = latest["result"] if latest else {}

    return {
        "layout_id": int(row["id"]),
        "tower": row["tower"] or "Unassigned Tower",
        "flat_number": row["flat_number"] or f'Layout {row["id"]}',
        "layout_type": row["layout_type"] or "—",
        "floor": row["floor"] or "—",
        "status": row["lifecycle_status"] or review_workflow.NOT_ANALYSED,
        "north": row["confirmed_north_orientation"] or "Unknown",
        "score": (
            float(row["overall_score"])
            if row["overall_score"] is not None
            else None
        ),
        "vastu_score": (
            float(row["vastu_score"])
            if row["vastu_score"] is not None
            else None
        ),
        "confidence": float(row["confidence_value"] or 0),
        "grade": row["grade"] or "—",
        "directions": final_directions,
        "strengths": _json(row["strengths_json"], []),
        "cautions": _json(row["cautions_json"], []),
        "findings": _json(row["findings_json"], []),
        "recommendations": _json(row["recommendations_json"], []),
        "knowledge": knowledge or {},
        "drawing_path": str(row["drawing_path"] or ""),
        "result_json_path": result_value,
    }


def project_export_data(project_id: int) -> dict[str, Any]:
    project = get_project(int(project_id))
    summary = consolidation_service.summary(int(project_id))
    layouts = [
        _layout_payload(row)
        for row in _latest_layout_rows(int(project_id))
    ]
    finalised = [
        item for item in layouts
        if item["status"] == review_workflow.FINALISED
        and item["score"] is not None
    ]

    tower_groups: dict[str, list[dict]] = defaultdict(list)
    issue_counter: Counter[str] = Counter()
    rule_counter: Counter[str] = Counter()

    for item in finalised:
        tower_groups[item["tower"]].append(item)
        for caution in item["cautions"]:
            issue_counter[str(caution)] += 1
        for finding in item["knowledge"].get("concerns", []):
            rule_counter[
                f'{finding.get("rule_id","")} · {finding.get("title","")}'
            ] += 1

    towers = []
    for tower, rows in tower_groups.items():
        scores = [row["score"] for row in rows if row["score"] is not None]
        towers.append(
            {
                "tower": tower,
                "finalised_layouts": len(rows),
                "average_score": (
                    round(sum(scores) / len(scores), 2)
                    if scores else None
                ),
                "best_flat": (
                    max(rows, key=lambda row: row["score"])["flat_number"]
                    if rows else "—"
                ),
                "lowest_flat": (
                    min(rows, key=lambda row: row["score"])["flat_number"]
                    if rows else "—"
                ),
            }
        )
    towers.sort(
        key=lambda item: item["average_score"] or -1,
        reverse=True,
    )

    return {
        "project": dict(project),
        "generated_at": datetime.now().astimezone().isoformat(),
        "summary": summary,
        "layouts": layouts,
        "finalised_layouts": finalised,
        "towers": towers,
        "common_issues": [
            {"issue": issue, "count": count}
            for issue, count in issue_counter.most_common(15)
        ],
        "common_knowledge_rules": [
            {"rule": rule, "count": count}
            for rule, count in rule_counter.most_common(15)
        ],
        "knowledge_profile": knowledge_service.get_profile(int(project_id)),
    }


def _styles():
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            "ReportTitle",
            parent=styles["Title"],
            fontSize=22,
            leading=27,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#285943"),
            spaceAfter=12,
        )
    )
    styles.add(
        ParagraphStyle(
            "Section",
            parent=styles["Heading2"],
            fontSize=14,
            leading=18,
            textColor=colors.HexColor("#285943"),
            spaceBefore=10,
            spaceAfter=7,
        )
    )
    styles.add(
        ParagraphStyle(
            "Small",
            parent=styles["BodyText"],
            fontSize=8,
            leading=10,
        )
    )
    styles.add(
        ParagraphStyle(
            "Disclaimer",
            parent=styles["BodyText"],
            fontSize=7.5,
            leading=10,
            textColor=colors.HexColor("#6B756F"),
        )
    )
    return styles


def _footer(canvas, document):
    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#6B756F"))
    canvas.drawString(
        18 * mm,
        10 * mm,
        "VastuAI - belief-based guidance; structural, legal and safety review take precedence.",
    )
    canvas.drawRightString(
        PAGE_WIDTH - 18 * mm,
        10 * mm,
        f"Page {document.page}",
    )
    canvas.restoreState()


def _table(data, widths=None, header=True, font_size=8):
    table = Table(data, colWidths=widths, repeatRows=1 if header else 0)
    commands = [
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B8C8C0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    if header:
        commands.extend(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#285943")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
        for row_index in range(1, len(data)):
            if row_index % 2 == 0:
                commands.append(("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#FAFCFA")))
    table.setStyle(TableStyle(commands))
    return table


def _metric_table(metrics: list[tuple[str, str]]):
    styles = _styles()
    rows = [[Paragraph(f"<b>{label}</b>", styles["Small"]), value] for label, value in metrics]
    table = Table(rows, colWidths=[55 * mm, 80 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EAF2EE")),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#B8C8C0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    return table


def _append_list(story, values: Iterable[str], styles, empty_text="None recorded."):
    values = list(values)
    if not values:
        story.append(Paragraph(empty_text, styles["BodyText"]))
        return
    for value in values:
        story.append(Paragraph(f"• {value}", styles["BodyText"]))


def _image_flowable(path_value: str, max_width=160 * mm, max_height=100 * mm):
    path = Path(path_value or "")
    if not path.exists() or not path.is_file():
        return None
    try:
        image = Image(str(path))
        width, height = image.imageWidth, image.imageHeight
        scale = min(max_width / width, max_height / height)
        image.drawWidth = width * scale
        image.drawHeight = height * scale
        return image
    except Exception:
        return None


def _knowledge_section(story, knowledge, styles):
    if not knowledge:
        story.append(
            Paragraph(
                "No Knowledge Engine assessment is stored for this result.",
                styles["BodyText"],
            )
        )
        return

    story.append(
        Paragraph(
            knowledge.get("reasoning_summary", ""),
            styles["BodyText"],
        )
    )
    findings = knowledge.get("findings", [])
    if findings:
        rows = [["Rule", "Category", "Finding", "Severity", "Confidence"]]
        for finding in findings:
            rows.append(
                [
                    finding.get("rule_id", ""),
                    finding.get("category", ""),
                    finding.get("title", ""),
                    finding.get("severity", ""),
                    f'{float(finding.get("combined_confidence", 0) or 0):.0%}',
                ]
            )
        story.append(
            _table(
                rows,
                widths=[22 * mm, 25 * mm, 72 * mm, 22 * mm, 24 * mm],
                font_size=7.2,
            )
        )

    actions = knowledge.get("priority_actions", [])
    if actions:
        story.append(Spacer(1, 5))
        story.append(Paragraph("Priority Knowledge Actions", styles["Heading3"]))
        for action in actions:
            story.append(
                Paragraph(
                    f'<b>{action.get("recommendation_id","")} - '
                    f'{action.get("title","Recommendation")}</b>',
                    styles["BodyText"],
                )
            )
            for step in action.get("actions", []):
                story.append(Paragraph(f"• {step}", styles["Small"]))


def individual_flat_pdf(project_id: int, layout_id: int) -> bytes:
    data = project_export_data(project_id)
    layout = next(
        (item for item in data["layouts"] if item["layout_id"] == int(layout_id)),
        None,
    )
    if not layout:
        raise ValueError("Layout not found.")

    buffer = io.BytesIO()
    styles = _styles()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f'Individual Flat Report - {layout["flat_number"]}',
        author="VastuAI",
    )
    story = [
        Spacer(1, 16 * mm),
        Paragraph("Individual Vastu Assessment", styles["ReportTitle"]),
        Paragraph(
            f'{data["project"]["name"]} · {layout["tower"]} · '
            f'{layout["flat_number"]}',
            styles["Heading2"],
        ),
        Spacer(1, 10 * mm),
        _metric_table(
            [
                ("Tower", layout["tower"]),
                ("Flat Number", layout["flat_number"]),
                ("Layout Type", layout["layout_type"]),
                ("Floor", layout["floor"]),
                ("Lifecycle Status", layout["status"]),
                ("Overall Score", (
                    f'{layout["score"]:.1f}/10'
                    if layout["score"] is not None else "—"
                )),
                ("Vastu Score", (
                    f'{layout["vastu_score"]:.1f}/10'
                    if layout["vastu_score"] is not None else "—"
                )),
                ("Confidence", f'{layout["confidence"]:.0%}'),
                ("Grade", layout["grade"]),
                ("Confirmed North", layout["north"]),
                ("Knowledge Profile", data["knowledge_profile"]),
            ]
        ),
        PageBreak(),
        Paragraph("Reviewed Layout", styles["Section"]),
    ]

    image = _image_flowable(layout["drawing_path"])
    if image:
        story.extend([image, Spacer(1, 5 * mm)])

    story.append(Paragraph("Final Reviewed Directions", styles["Section"]))
    labels = dict(review_workflow.ROOM_FIELDS)
    rows = [["Room", "Direction"]]
    for field, label in review_workflow.ROOM_FIELDS:
        rows.append([label, layout["directions"].get(field, "Unknown")])
    story.append(_table(rows, widths=[75 * mm, 75 * mm]))

    story.append(Paragraph("Strengths", styles["Section"]))
    _append_list(story, layout["strengths"], styles)

    story.append(Paragraph("Needs Attention", styles["Section"]))
    _append_list(story, layout["cautions"], styles)

    story.append(Paragraph("Professional Recommendations", styles["Section"]))
    if layout["recommendations"]:
        for recommendation in layout["recommendations"]:
            if isinstance(recommendation, dict):
                title = (
                    recommendation.get("area")
                    or recommendation.get("title")
                    or "Recommendation"
                )
                story.append(Paragraph(f"<b>{title}</b>", styles["BodyText"]))
                for key in (
                    "finding",
                    "why_it_matters",
                    "practical_action",
                    "structural_option",
                ):
                    if recommendation.get(key):
                        story.append(
                            Paragraph(
                                f'<b>{key.replace("_"," ").title()}:</b> '
                                f'{recommendation[key]}',
                                styles["Small"],
                            )
                        )
            else:
                story.append(Paragraph(f"• {recommendation}", styles["BodyText"]))
    else:
        story.append(Paragraph("No recommendations recorded.", styles["BodyText"]))

    story.append(PageBreak())
    story.append(Paragraph("Knowledge Engine Assessment", styles["Section"]))
    _knowledge_section(story, layout["knowledge"], styles)
    story.append(Spacer(1, 8 * mm))
    story.append(
        Paragraph(
            "This report is generated from the final reviewed room directions "
            "and the locally stored Knowledge Objects. Vastu and numerology are "
            "belief-based systems. This report is not structural, engineering, "
            "legal, financial or medical advice.",
            styles["Disclaimer"],
        )
    )

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def _tower_layouts(data, tower: str):
    return [
        item for item in data["finalised_layouts"]
        if item["tower"] == tower
    ]


def tower_pdf(project_id: int, tower: str) -> bytes:
    data = project_export_data(project_id)
    layouts = _tower_layouts(data, tower)
    buffer = io.BytesIO()
    styles = _styles()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=16 * mm,
        bottomMargin=18 * mm,
        title=f"Tower Vastu Report - {tower}",
        author="VastuAI",
    )
    scores = [item["score"] for item in layouts if item["score"] is not None]
    avg = round(sum(scores) / len(scores), 2) if scores else None

    story = [
        Paragraph("Tower Vastu Report", styles["ReportTitle"]),
        Paragraph(
            f'{data["project"]["name"]} · {tower}',
            styles["Heading2"],
        ),
        _metric_table(
            [
                ("Tower", tower),
                ("Finalised Flats", str(len(layouts))),
                ("Average Score", f"{avg:.1f}/10" if avg is not None else "—"),
                ("Knowledge Profile", data["knowledge_profile"]),
                ("Generated", data["generated_at"]),
            ]
        ),
        Spacer(1, 6 * mm),
        Paragraph("Flat Ranking", styles["Section"]),
    ]
    rows = [["Rank", "Flat", "Type", "Floor", "Score", "Confidence", "Grade"]]
    for index, item in enumerate(
        sorted(layouts, key=lambda x: x["score"], reverse=True),
        start=1,
    ):
        rows.append(
            [
                index,
                item["flat_number"],
                item["layout_type"],
                item["floor"],
                f'{item["score"]:.1f}',
                f'{item["confidence"]:.0%}',
                item["grade"],
            ]
        )
    story.append(
        _table(
            rows,
            widths=[12*mm, 30*mm, 30*mm, 20*mm, 18*mm, 24*mm, 20*mm],
            font_size=7.5,
        )
    )

    issue_counter = Counter()
    rule_counter = Counter()
    for item in layouts:
        for caution in item["cautions"]:
            issue_counter[str(caution)] += 1
        for finding in item["knowledge"].get("concerns", []):
            rule_counter[
                f'{finding.get("rule_id","")} · {finding.get("title","")}'
            ] += 1

    story.append(Paragraph("Recurring Professional Concerns", styles["Section"]))
    issue_rows = [["Concern", "Flats"]]
    issue_rows.extend(
        [[name, count] for name, count in issue_counter.most_common(12)]
    )
    story.append(_table(issue_rows, widths=[140*mm, 22*mm], font_size=7.5))

    story.append(Paragraph("Recurring Knowledge Rules", styles["Section"]))
    rule_rows = [["Rule", "Flats"]]
    rule_rows.extend(
        [[name, count] for name, count in rule_counter.most_common(12)]
    )
    story.append(_table(rule_rows, widths=[140*mm, 22*mm], font_size=7.5))

    story.append(Paragraph("Priority Recommendations", styles["Section"]))
    action_counter = Counter()
    action_lookup = {}
    for item in layouts:
        for action in item["knowledge"].get("priority_actions", []):
            rec_id = action.get("recommendation_id", "")
            if rec_id:
                action_counter[rec_id] += 1
                action_lookup[rec_id] = action
    if action_counter:
        for rec_id, count in action_counter.most_common(10):
            action = action_lookup[rec_id]
            story.append(
                Paragraph(
                    f'<b>{rec_id} · {action.get("title","Recommendation")}</b> '
                    f'({count} flat(s))',
                    styles["BodyText"],
                )
            )
    else:
        story.append(Paragraph("No recurring Knowledge actions.", styles["BodyText"]))

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def building_pdf(project_id: int) -> bytes:
    data = project_export_data(project_id)
    summary = data["summary"]
    buffer = io.BytesIO()
    styles = _styles()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=16 * mm,
        bottomMargin=18 * mm,
        title=f'Building Vastu Report - {data["project"]["name"]}',
        author="VastuAI",
    )
    score = summary["building_score"]
    story = [
        Paragraph("VastuAI Builder Project Report", styles["ReportTitle"]),
        Paragraph(data["project"]["name"], styles["Heading2"]),
        _metric_table(
            [
                ("Total Layouts", str(summary["total"])),
                ("Finalised Layouts", str(summary["finalised"])),
                ("Finalisation Coverage", f'{summary["coverage"]:.0%}'),
                ("Building Score", (
                    f"{score:.1f}/10" if score is not None else "—"
                )),
                ("Knowledge Profile", data["knowledge_profile"]),
                ("Generated", data["generated_at"]),
            ]
        ),
    ]

    if summary["coverage"] < 0.5:
        story.append(
            Paragraph(
                "The building score is provisional because fewer than 50% "
                "of layouts are finalised.",
                styles["Disclaimer"],
            )
        )

    story.append(Paragraph("Tower Comparison", styles["Section"]))
    tower_rows = [["Rank", "Tower", "Finalised Flats", "Average", "Best Flat", "Lowest Flat"]]
    for index, item in enumerate(data["towers"], start=1):
        tower_rows.append(
            [
                index,
                item["tower"],
                item["finalised_layouts"],
                (
                    f'{item["average_score"]:.1f}'
                    if item["average_score"] is not None else "—"
                ),
                item["best_flat"],
                item["lowest_flat"],
            ]
        )
    story.append(
        _table(
            tower_rows,
            widths=[12*mm, 33*mm, 25*mm, 22*mm, 35*mm, 35*mm],
            font_size=7.5,
        )
    )

    story.append(Paragraph("Finalised Flat Ranking", styles["Section"]))
    flat_rows = [["Rank", "Tower", "Flat", "Type", "Floor", "Score", "Grade"]]
    for index, item in enumerate(
        sorted(
            data["finalised_layouts"],
            key=lambda row: row["score"],
            reverse=True,
        ),
        start=1,
    ):
        flat_rows.append(
            [
                index,
                item["tower"],
                item["flat_number"],
                item["layout_type"],
                item["floor"],
                f'{item["score"]:.1f}',
                item["grade"],
            ]
        )
    story.append(
        _table(
            flat_rows,
            widths=[10*mm, 28*mm, 28*mm, 28*mm, 20*mm, 18*mm, 20*mm],
            font_size=7.2,
        )
    )

    story.append(Paragraph("Common Professional Concerns", styles["Section"]))
    concern_rows = [["Concern", "Occurrences"]]
    concern_rows.extend(
        [[item["issue"], item["count"]] for item in data["common_issues"]]
    )
    story.append(
        _table(concern_rows, widths=[145*mm, 22*mm], font_size=7.5)
    )

    story.append(Paragraph("Common Knowledge Rules", styles["Section"]))
    rule_rows = [["Rule", "Occurrences"]]
    rule_rows.extend(
        [[item["rule"], item["count"]]
         for item in data["common_knowledge_rules"]]
    )
    story.append(_table(rule_rows, widths=[145*mm, 22*mm], font_size=7.5))

    story.append(PageBreak())
    story.append(Paragraph("Flat Appendix", styles["Section"]))
    for item in sorted(
        data["finalised_layouts"],
        key=lambda row: (row["tower"], row["flat_number"]),
    ):
        block = [
            Paragraph(
                f'<b>{item["tower"]} · {item["flat_number"]}</b>',
                styles["Heading3"],
            ),
            Paragraph(
                f'Score: {item["score"]:.1f}/10 · Grade: {item["grade"]} · '
                f'Confidence: {item["confidence"]:.0%} · North: {item["north"]}',
                styles["Small"],
            ),
        ]
        if item["cautions"]:
            block.append(
                Paragraph(
                    "Attention: " + "; ".join(map(str, item["cautions"][:4])),
                    styles["Small"],
                )
            )
        story.extend([KeepTogether(block), Spacer(1, 4 * mm)])

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def dashboard_pdf(project_id: int) -> bytes:
    data = project_export_data(project_id)
    buffer = io.BytesIO()
    styles = _styles()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=15 * mm,
        title=f'Dashboard Export - {data["project"]["name"]}',
        author="VastuAI",
    )
    summary = data["summary"]
    story = [
        Paragraph("VastuAI Dashboard Export", styles["ReportTitle"]),
        Paragraph(data["project"]["name"], styles["Heading2"]),
        _metric_table(
            [
                ("Total Layouts", str(summary["total"])),
                ("Finalised", str(summary["finalised"])),
                ("Coverage", f'{summary["coverage"]:.0%}'),
                ("Building Score", (
                    f'{summary["building_score"]:.1f}/10'
                    if summary["building_score"] is not None else "—"
                )),
            ]
        ),
        Paragraph("Tower Ranking", styles["Section"]),
    ]
    tower_rows = [["Tower", "Finalised Layouts", "Average Score", "Best", "Lowest"]]
    tower_rows.extend(
        [
            [
                row["tower"],
                row["finalised_layouts"],
                row["average_score"],
                row["best_flat"],
                row["lowest_flat"],
            ]
            for row in data["towers"]
        ]
    )
    story.append(
        _table(
            tower_rows,
            widths=[48*mm, 35*mm, 35*mm, 45*mm, 45*mm],
            font_size=8,
        )
    )
    story.append(Paragraph("Top Finalised Flats", styles["Section"]))
    flat_rows = [["Rank", "Tower", "Flat", "Type", "Floor", "Score", "Confidence", "Grade"]]
    for i, row in enumerate(
        sorted(data["finalised_layouts"], key=lambda x: x["score"], reverse=True)[:25],
        start=1,
    ):
        flat_rows.append(
            [
                i,
                row["tower"],
                row["flat_number"],
                row["layout_type"],
                row["floor"],
                row["score"],
                f'{row["confidence"]:.0%}',
                row["grade"],
            ]
        )
    story.append(_table(flat_rows, font_size=7.2))
    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


def csv_exports(project_id: int) -> dict[str, bytes]:
    data = project_export_data(project_id)
    layouts = pd.DataFrame(data["layouts"])
    if not layouts.empty:
        layouts = layouts.drop(
            columns=[
                "directions",
                "strengths",
                "cautions",
                "findings",
                "recommendations",
                "knowledge",
            ],
            errors="ignore",
        )
    towers = pd.DataFrame(data["towers"])
    issues = pd.DataFrame(data["common_issues"])
    rules = pd.DataFrame(data["common_knowledge_rules"])
    return {
        "layouts.csv": layouts.to_csv(index=False).encode("utf-8-sig"),
        "towers.csv": towers.to_csv(index=False).encode("utf-8-sig"),
        "common_issues.csv": issues.to_csv(index=False).encode("utf-8-sig"),
        "knowledge_rules.csv": rules.to_csv(index=False).encode("utf-8-sig"),
    }


def excel_export(project_id: int) -> bytes:
    data = project_export_data(project_id)
    workbook = Workbook()
    workbook.remove(workbook.active)

    def add_sheet(title, headers, rows):
        sheet = workbook.create_sheet(title)
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        fill = PatternFill("solid", fgColor="DDF3E5")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="365F48")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        for column in range(1, len(headers) + 1):
            max_length = max(
                len(str(sheet.cell(row=row, column=column).value or ""))
                for row in range(1, sheet.max_row + 1)
            )
            sheet.column_dimensions[get_column_letter(column)].width = min(
                max(max_length + 2, 12),
                45,
            )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    summary = data["summary"]
    add_sheet(
        "Project Summary",
        ["Metric", "Value"],
        [
            ["Project", data["project"]["name"]],
            ["Generated", data["generated_at"]],
            ["Knowledge Profile", data["knowledge_profile"]],
            ["Total Layouts", summary["total"]],
            ["Finalised", summary["finalised"]],
            ["Coverage", summary["coverage"]],
            ["Building Score", summary["building_score"]],
        ],
    )

    add_sheet(
        "Layouts",
        [
            "Tower", "Flat Number", "Layout Type", "Floor", "Status",
            "Score", "Vastu Score", "Confidence", "Grade", "North",
        ],
        [
            [
                row["tower"], row["flat_number"], row["layout_type"],
                row["floor"], row["status"], row["score"],
                row["vastu_score"], row["confidence"], row["grade"],
                row["north"],
            ]
            for row in data["layouts"]
        ],
    )

    add_sheet(
        "Tower Ranking",
        ["Tower", "Finalised Layouts", "Average Score", "Best Flat", "Lowest Flat"],
        [
            [
                row["tower"], row["finalised_layouts"], row["average_score"],
                row["best_flat"], row["lowest_flat"],
            ]
            for row in data["towers"]
        ],
    )

    add_sheet(
        "Common Issues",
        ["Issue", "Count"],
        [[row["issue"], row["count"]] for row in data["common_issues"]],
    )

    add_sheet(
        "Knowledge Rules",
        ["Rule", "Count"],
        [[row["rule"], row["count"]]
         for row in data["common_knowledge_rules"]],
    )

    findings_rows = []
    for layout in data["finalised_layouts"]:
        for finding in layout["knowledge"].get("findings", []):
            findings_rows.append(
                [
                    layout["tower"],
                    layout["flat_number"],
                    finding.get("rule_id"),
                    finding.get("category"),
                    finding.get("title"),
                    finding.get("observed_value"),
                    finding.get("polarity"),
                    finding.get("severity"),
                    finding.get("combined_confidence"),
                ]
            )
    add_sheet(
        "Knowledge Findings",
        [
            "Tower", "Flat", "Rule ID", "Category", "Finding",
            "Direction", "Polarity", "Severity", "Confidence",
        ],
        findings_rows,
    )

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
